import random
from typing import List, Tuple
from collections import defaultdict

import networkx as nx
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def build_schedule_flow(
    names: List[str],
    weeks: int ,
    max_pair_repeat: int ,
    max_attempts: int ,
    seed: int | None = None,
) -> List[List[Tuple[int, int]]]:

    N = len(names)
    if weeks < 1:
        raise ValueError("weeks ต้อง ≥ 1")
    if N < weeks + 1:
        raise ValueError(f"ต้องมีคน ≥ weeks+1. ตอนนี้คน={N}, weeks={weeks}")

    rng = random.Random(seed)

    def try_once() -> List[List[Tuple[int, int]]] | None:
        remain = [[0]*N for _ in range(N)]
        for i in range(N):
            for j in range(N):
                remain[i][j] = 0 if i == j else max_pair_repeat # ตรวจคนเดิมได้มากสุด 2 และห้ามตรวจตัวเอง

        weekly_pairs: List[List[Tuple[int, int]]] = []

        for w in range(weeks):
            reviewers = list(range(N))
            targets = list(range(N))
            rng.shuffle(reviewers)
            rng.shuffle(targets)

            G = nx.DiGraph()
            SRC, SNK = "S", "T"
            
            for i in reviewers: # กำหนดให้ตรวจได้ 2 คน
                G.add_edge(SRC, f"r{i}", capacity=2)
            
            for i in reviewers: # ตรวจใครได้บ้าง คนเดิมได้แค่ 1 ครั้งต่อสัปดาห์
                for j in targets:
                    if i == j:
                        continue
                    if remain[i][j] > 0:
                        G.add_edge(f"r{i}", f"t{j}", capacity=1)
                       

            for j in targets:  # กำหนดให้ถูกตรวจได้ 2 ครั้ง
                G.add_edge(f"t{j}", SNK, capacity=2)

            flow_value, flow_dict = nx.maximum_flow(G, SRC, SNK)
            if flow_value != 2 * N: # ต้องมีโฟลว์ = 2N ถึงจะเป๊ะ
                return None  

            # แปลงผล flow เป็นลิสต์เป้าหมายสองคนต่อ reviewer i
            week_assign = [[] for _ in range(N)]
            for i in range(N):
                edges = flow_dict.get(f"r{i}", {})
                bag = []
                # iterate แบบสุ่มเพื่อหลีกเลี่ยง pattern เดิม
                out_items = list(edges.items())
                rng.shuffle(out_items)
                for v, f in out_items:
                    if f <= 0 or not v.startswith("t"):
                        continue
                    j = int(v[1:])
                    bag.extend([j] * int(f))
                if len(bag) != 2:
                    # ควรเท่ากับ 2 เสมอถ้า flow=2N
                    return None
                week_assign[i] = bag

            # commit สัปดาห์นี้
            weekly_pairs.append([(week_assign[i][0], week_assign[i][1]) for i in range(N)])
            for i in range(N):
                for j in week_assign[i]:
                    remain[i][j] -= 1
                    if remain[i][j] < 0:
                        return None  # safety

        return weekly_pairs

    for _ in range(max_attempts):
        ans = try_once()
        if ans is not None:
            return ans

    raise RuntimeError("จัดตารางไม่สำเร็จภายใต้ข้อจำกัด ลองเพิ่มจำนวนคนหรือเพิ่ม max_pair_repeat หรือเปลี่ยน weeks/seed")
def export_excel_with_sendingtime(names, weekly_pairs, out_path):
    N = len(names); W = len(weekly_pairs)
    wb = Workbook()

    # ---------- Build incoming map: for each week & target -> [(reviewer, slot1or2)] ----------
    incoming = [[[] for _ in range(N)] for _ in range(W)]
    for w in range(W):
        for i in range(N):
            t1, t2 = weekly_pairs[w][i]
            incoming[w][t1].append((i, 1))
            incoming[w][t2].append((i, 2))

    # ---------- SendingTime ----------
    st = wb.active; st.title = "SendingTime"
    st.cell(row=1, column=1, value="Week")
    for w in range(W):
        st.cell(row=2+w, column=1, value=w+1)

    col_map = {}
    col = 2
    for name in names:
        st.cell(row=1, column=col,   value=f"{name} Time");        time_col = col;   col += 1
        st.cell(row=1, column=col,   value=f"{name} EvidenceURL"); url_col  = col;   col += 1
        st.cell(row=1, column=col,   value=f"{name} got checked"); status_col = col; col += 1
        col_map[name] = (time_col, url_col, status_col)

    for c in range(1, col):
        st.column_dimensions[get_column_letter(c)].width = 20

    # ---------- Individual sheets (add Process) ----------
    for i, name in enumerate(names):
        ws = wb.create_sheet(title=(name if len(name)<=31 else name[:31]))
        ws.append(["Week","Reviewer","Target #1","T1 Time","Process","Target #2","T2 Time","Process"])

        reviewer = names[i]
        for w in range(W):
            t1, t2 = weekly_pairs[w][i]
            # addresses in SendingTime
            st_row = 2 + w
            t1_time_c, t1_url_c, _ = col_map[names[t1]]
            t2_time_c, t2_url_c, _ = col_map[names[t2]]
            t1_time_addr = f"{get_column_letter(t1_time_c)}{st_row}"
            t2_time_addr = f"{get_column_letter(t2_time_c)}{st_row}"

            f_t1_time = f"=SendingTime!{t1_time_addr}"
            f_t2_time = f"=SendingTime!{t2_time_addr}"


            # Process default FALSE (จะไปแปลงเป็น checkbox ใน Google Sheets ได้ทีหลัง)
            ws.append([w+1, reviewer, names[t1], f_t1_time, False, names[t2], f_t2_time, False])

        for c in range(1, 9):
            ws.column_dimensions[get_column_letter(c)].width = 18

    # ---------- Fill "got checked" = AND(Process of both reviewers this week) in SendingTime ----------
    # row in reviewer-sheet = week+1
    for w in range(W):
        st_row = 2 + w
        for j, name in enumerate(names):
            _, _, status_col = col_map[name]
            pair = incoming[w][j]  # list of two: (reviewer_idx, slot)
            if len(pair) != 2:
                # safety, should always be 2
                continue
            (r1, s1), (r2, s2) = pair
            col1 = "E" if s1 == 1 else "H"  # Process col in reviewer sheet
            col2 = "E" if s2 == 1 else "H"
            r1_sheet = names[r1] if len(names[r1])<=31 else names[r1][:31]
            r2_sheet = names[r2] if len(names[r2])<=31 else names[r2][:31]
            r = str(1 + w + 1)  # header row=1, data start=2 ⇒ week row = w+2
            formula = f"=AND('{r1_sheet}'!{col1}{r}, '{r2_sheet}'!{col2}{r})"
            st.cell(row=st_row, column=status_col, value=formula)

    wb.save(out_path)

if __name__ == "__main__":
    names = [
        "Lucas","Knot","Margin","JingJang","Tar",
        "Chin","Pat","Junior","Wave","Peace"
    ]
    weeks = 9
    max_pair_repeat = 2
    seed = 42  # เปลี่ยนได้ หรือ None

    weekly_pairs = build_schedule_flow(
        names,
        weeks=weeks,
        max_pair_repeat=max_pair_repeat,
        max_attempts=50,
        seed=seed,
    )
    export_excel_with_sendingtime(names, weekly_pairs, "review_schedule.xlsx")
    print("OK -> review_schedule.xlsx")
